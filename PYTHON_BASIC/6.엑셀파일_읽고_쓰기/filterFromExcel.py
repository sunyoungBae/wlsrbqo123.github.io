import openpyxl
'''
wb = openpyxl.load_workbook('./6.엑셀파일_읽고_쓰기/test.xlsx')
sheet = wb['Sheet1']

# 엑셀 파일의 특정 셀 내용 읽어오기
print(sheet.max_column, sheet.max_row)
print(sheet.cell(row=1, column=1).value) # 각 셀의 값
print(sheet.cell(row=2, column=1).value)
print('-' * 10)

# 엑셀 파일의 모든 셀 내용 읽어오기
print("data from 2 row")
for row in sheet.iter_rows(min_row=2): # 두 번째 행부터 탐색
    for cell in row: # 열 단위 데이터 탐색
        print(cell.value)
    print('-' * 10)

# 엑셀 파일에서 지정한 셀 내용 읽어오기
print("data from A2 : C3")
cells = sheet['A2':'C3']
for row in cells:
    for cell in row:
        print(cell.value)
    print('-' * 10)

wb.close()
'''
'''
#  회원정보를 엑셀에 저장
wb = openpyxl.Workbook()
sheet = wb.active # 현재 활성화된 워크 시트 객체를 가져옴
sheet.title = '회원정보'

# 표 헤더 컬럼 저장
header_titles = ['이름', '전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1, column=idx+1, value=title)

# 표 내용 저장
members = [('kei', '010-1234-1234'), ('hong', '010-4321-1234')]
row_num = 2
for r, member in enumerate(members):
    for c, v in enumerate(member):
        sheet.cell(row=row_num+r, column=c+1, value=v)

wb.save('./6.엑셀파일_읽고_쓰기/members.xlsx')
wb.close()
'''